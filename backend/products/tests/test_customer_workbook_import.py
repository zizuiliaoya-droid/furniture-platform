"""客户横向多 Sheet 配置模板回归测试。"""
from io import BytesIO

import openpyxl
import pytest

from products.models import Product, ProductConfigPreset, ProductPriceMatrix
from products.services import CustomerWorkbookImportService


def workbook_bytes(workbook):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def customer_workbook():
    workbook = openpyxl.Workbook()
    info = workbook.active
    info.title = '产品信息'
    info.append(['配置Sheet', '产品编号', '产品名称', '一级分类', '二级分类'])
    info.append(['椅配置', 'CHAIR-01', '测试椅', 'SEATING', 'TASK_CHAIR'])
    info.append(['桌配置', 'DESK-01', '测试桌', 'DESKS_WORKSTATIONS', 'FIXED_DESK'])
    chair = workbook.create_sheet('椅配置')
    chair.append(['坐垫饰面', '布（二级）', '皮（二级）', '背饰面', '布（二级）', '网（二级）'])
    chair.append(['布', 'P1', 'L1', '网', 'P2', 'M1'])
    chair.append(['皮', 'P3', 'L2', '布', 'P4', 'M2'])
    desk = workbook.create_sheet('桌配置')
    desk.append(['屏风材质', '布', 'PET', '钢', '隐藏项'])
    desk.append(['布', 'B1', 'PET1', 'S1', '忽略'])
    desk.column_dimensions['E'].hidden = True
    prices = workbook.create_sheet('组合价格')
    prices.append(['配置Sheet', '组合编号', '组合名称', '是否默认', '最终价格',
                   '坐垫饰面', '坐垫饰面_布', '背饰面', '背饰面_网'])
    prices.append(['椅配置', 'STD', '标准配置', '是', 3580, '布', 'P1', '网', 'M1'])
    return workbook


@pytest.mark.django_db
class TestCustomerWorkbookImport:
    def test_context_children_price_and_default_preset(self, admin_user):
        parsed = CustomerWorkbookImportService.parse(workbook_bytes(customer_workbook()))
        assert parsed['errors'] == []
        chair_dims = parsed['products']['椅配置']['dimensions']
        seat_fabric = next(dim for dim in chair_dims.values() if dim['label'] == '坐垫饰面-布')
        back_fabric = next(dim for dim in chair_dims.values() if dim['label'] == '背饰面-布')
        assert seat_fabric['key'] != back_fabric['key']
        assert seat_fabric['parent_dimension'].endswith('=布')
        desk_dims = parsed['products']['桌配置']['dimensions']
        for label in ('屏风材质-布', '屏风材质-PET', '屏风材质-钢'):
            assert next(dim for dim in desk_dims.values() if dim['label'] == label)['parent_dimension']
        assert all(dim['label'] != '隐藏项' for dim in desk_dims.values())
        assert any('隐藏列' in warning for warning in parsed['warnings'])

        result = CustomerWorkbookImportService.execute_import(parsed, admin_user)
        assert result['created'] == 2
        chair = Product.objects.get(code='CHAIR-01')
        assert ProductPriceMatrix.objects.filter(product=chair).count() == 1
        preset = ProductConfigPreset.objects.get(product=chair, is_default=True)
        assert preset.selections
        assert chair.min_price == 3580

    def test_idempotent_update(self, admin_user):
        first = CustomerWorkbookImportService.parse(workbook_bytes(customer_workbook()))
        CustomerWorkbookImportService.execute_import(first, admin_user)
        second = CustomerWorkbookImportService.parse(workbook_bytes(customer_workbook()))
        result = CustomerWorkbookImportService.execute_import(second, admin_user)
        assert result['updated'] == 2
        assert Product.objects.filter(code='CHAIR-01').count() == 1

    def test_options_only_update_preserves_existing_prices_and_default(self, admin_user):
        initial = CustomerWorkbookImportService.parse(workbook_bytes(customer_workbook()))
        CustomerWorkbookImportService.execute_import(initial, admin_user)
        chair = Product.objects.get(code='CHAIR-01')
        original_min_price = chair.min_price
        original_matrix_count = chair.price_matrix.count()
        original_preset_count = chair.config_presets.count()

        options_only = customer_workbook()
        options_only.remove(options_only['组合价格'])
        parsed = CustomerWorkbookImportService.parse(workbook_bytes(options_only))
        assert parsed['errors'] == []
        assert any('暂无组合价格' in warning for warning in parsed['warnings'])
        CustomerWorkbookImportService.execute_import(parsed, admin_user)

        chair.refresh_from_db()
        assert chair.price_matrix.count() == original_matrix_count == 1
        assert chair.config_presets.count() == original_preset_count == 1
        assert chair.min_price == original_min_price == 3580

    def test_duplicate_product_code_and_name_are_preview_errors(self):
        workbook = customer_workbook()
        workbook['产品信息']['B3'] = 'CHAIR-01'
        workbook['产品信息']['C3'] = '测试椅'
        parsed = CustomerWorkbookImportService.parse(workbook_bytes(workbook))
        assert any('产品编号' in error and '重复' in error for error in parsed['errors'])
        assert any('产品名称' in error and '重复' in error for error in parsed['errors'])

    def test_duplicate_matrix_signature_is_preview_error(self):
        workbook = customer_workbook()
        workbook['组合价格'].append(
            ['椅配置', 'DUP', '重复组合', '否', 4000, '布', 'P1', '网', 'M1'])
        parsed = CustomerWorkbookImportService.parse(workbook_bytes(workbook))
        assert any('配置组合' in error and '重复' in error for error in parsed['errors'])