"""Product management services."""
import hashlib
import json
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.conf import settings
from django.db import transaction
from django.db.models import Q
from django.utils.text import slugify

from common.file_storage import FileStorageService
from .models import (
    Brand, Category, Product, ProductConfig, ProductConfigDimension,
    ProductConfigPreset, ProductImage, ProductPriceMatrix, ProductPriceRule,
)


class CategoryService:
    @staticmethod
    def get_tree(dimension: str):
        return Category.objects.filter(
            dimension=dimension, parent__isnull=True
        ).order_by('sort_order', 'id')

    @staticmethod
    def reorder(items: list):
        for item in items:
            Category.objects.filter(pk=item['id']).update(sort_order=item['sort_order'])


class ProductImageService:
    @staticmethod
    def upload_images(product: Product, files: list) -> list:
        created = []
        uploaded_paths = []
        try:
            for f in files:
                if f.size > settings.MAX_IMAGE_SIZE:
                    raise ValueError(f'图片 {f.name} 超过大小限制')
                path = FileStorageService.upload(f, 'products')
                uploaded_paths.append(path)
                thumbs = FileStorageService.generate_thumbnails(path)
                img = ProductImage.objects.create(
                    product=product,
                    image_path=path,
                    thumbnail_path=thumbs,
                    sort_order=product.images.count(),
                    is_cover=not product.images.exists(),
                )
                created.append(img)
            return created
        except Exception:
            # 数据库事务回滚不会清除已经写入磁盘的文件，必须主动补偿。
            for path in uploaded_paths:
                FileStorageService.delete_with_thumbnails(path)
            raise

    @staticmethod
    def delete_image(image: ProductImage):
        FileStorageService.delete_with_thumbnails(image.image_path)
        image.delete()

    @staticmethod
    def set_cover(image: ProductImage):
        ProductImage.objects.filter(product=image.product, is_cover=True).update(is_cover=False)
        image.is_cover = True
        image.save(update_fields=['is_cover'])


class ProductImportService:
    """产品批量导入（旧 Excel 导入，保留兼容）"""
    REQUIRED_HEADERS = ['名称', '产地']
    ORIGIN_MAP = {'进口': 'IMPORT', '国产': 'DOMESTIC'}

    @staticmethod
    def parse_excel(file) -> dict:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            return {'success_count': 0, 'failed_count': 0, 'preview': []}
        headers = [str(h).strip() if h else '' for h in rows[0]]
        results = []
        seen_codes = set()
        for i, row in enumerate(rows[1:], start=2):
            data = dict(zip(headers, row))
            errors = []
            name = str(data.get('名称', '') or '').strip()
            if not name:
                errors.append('名称为必填项')
            origin_raw = str(data.get('产地', '') or '').strip()
            origin = ProductImportService.ORIGIN_MAP.get(origin_raw)
            if not origin:
                errors.append(f'产地无效: {origin_raw}')
            code = str(data.get('编号', '') or '').strip() or None
            if code:
                if code in seen_codes or Product.objects.filter(code=code).exists():
                    errors.append(f'编号重复: {code}')
                seen_codes.add(code)
            results.append({
                'row': i, 'name': name, 'code': code, 'origin': origin,
                'description': str(data.get('描述', '') or ''),
                'min_price': data.get('最低售价'),
                'errors': errors,
            })
        success = [r for r in results if not r['errors']]
        failed = [r for r in results if r['errors']]
        return {'success_count': len(success), 'failed_count': len(failed), 'preview': results, 'parsed_data': success}

    @staticmethod
    @transaction.atomic
    def execute_import(parsed_data: list, user) -> int:
        count = 0
        for item in parsed_data:
            Product.objects.create(
                name=item['name'], code=item['code'], origin=item['origin'],
                description=item['description'],
                min_price=item['min_price'] if item['min_price'] else None,
                created_by=user,
            )
            count += 1
        return count

    @staticmethod
    def generate_template():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '产品导入模板'
        ws.append(['名称', '编号', '产地', '描述', '最低售价'])
        ws.append(['示例产品', 'P001', '进口', '产品描述', 1000])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


# ─── 价格计算服务 ─────────────────────────────────────────────────────────────

class PriceCalculationService:
    """产品配置实时计算指导价"""

    @staticmethod
    def _option_keys(dim: 'ProductConfigDimension') -> list:
        """提取一个维度允许的所有 option key"""
        keys = []
        for opt in dim.options or []:
            if isinstance(opt, dict):
                keys.append(opt.get('key', ''))
            else:
                keys.append(str(opt))
        return [k for k in keys if k]

    @staticmethod
    def _is_visible(dim: 'ProductConfigDimension', selections: dict) -> bool:
        parent = (dim.parent_dimension or '').strip()
        if not parent:
            return True
        parent_parts = parent.split('=', 1)
        parent_key = parent_parts[0].strip()
        parent_value = selections.get(parent_key)
        if parent_value in (None, ''):
            return False
        if len(parent_parts) == 2 and str(parent_value) != parent_parts[1].strip():
            return False
        return True

    @staticmethod
    def calculate(product: Product, selections: dict) -> dict:
        """校验当前可见配置并按 Matrix 或 Rule 计算价格。"""
        selections = ProductPriceMatrix.normalize_selections(selections)
        all_dims = list(ProductConfigDimension.objects.filter(product=product))
        dim_by_key = {d.dimension_key: d for d in all_dims}

        # 无配置维度的固定价产品也能进入报价单。
        if not all_dims:
            fixed_price = product.base_price if product.base_price is not None else product.min_price
            if fixed_price is None:
                return {
                    'valid': False, 'price': None, 'missing_dimensions': [],
                    'invalid_selections': [], 'breakdown': {}, 'reason': '产品尚未配置价格',
                }
            return {
                'valid': True, 'price': fixed_price, 'missing_dimensions': [],
                'invalid_selections': [], 'breakdown': {}, 'reason': '',
            }

        # 只校验当前父条件下可见的必填维度；不可见子维度不应误报缺失。
        visible_dims = [d for d in all_dims if PriceCalculationService._is_visible(d, selections)]
        missing = [
            d.dimension_key for d in visible_dims
            if d.is_required and d.dimension_key not in selections
        ]
        if missing:
            return {
                'valid': False,
                'price': None,
                'missing_dimensions': missing,
                'invalid_selections': [],
                'breakdown': {},
                'reason': '缺少必填配置维度',
            }

        # 2. 校验每个 selection 的维度合法性 + 选项归属
        invalid: list = []
        for key, value in selections.items():
            dim = dim_by_key.get(key)
            if dim is None:
                invalid.append({
                    'dimension_key': key,
                    'option_key': value,
                    'reason': f'维度 {key} 不属于该产品',
                })
                continue
            allowed = PriceCalculationService._option_keys(dim)
            if allowed and value not in allowed:
                invalid.append({
                    'dimension_key': key,
                    'option_key': value,
                    'reason': f'选项 {value} 不在维度 {key} 的允许列表 {allowed}',
                })
        if invalid:
            return {
                'valid': False,
                'price': None,
                'missing_dimensions': [],
                'invalid_selections': invalid,
                'breakdown': {},
                'reason': '存在非法选项',
            }

        # 3. 校验级联约束
        # parent_dimension 格式：
        #   "parent_key"               — 仅要求父维度已选（任意值）
        #   "parent_key=parent_value"  — 要求父维度已选且值等于 parent_value
        for dim in all_dims:
            if not dim.parent_dimension:
                continue
            if dim.dimension_key not in selections:
                # 子维度可选；用户没选则不约束
                continue
            parent_parts = dim.parent_dimension.split('=', 1)
            parent_key = parent_parts[0].strip()
            required_parent_value = parent_parts[1].strip() if len(parent_parts) == 2 else None

            if parent_key not in selections:
                return {
                    'valid': False,
                    'price': None,
                    'missing_dimensions': [parent_key],
                    'invalid_selections': [],
                    'breakdown': {},
                    'reason': f'维度 {dim.dimension_key} 依赖 {parent_key}，请先选择',
                }
            if required_parent_value is not None and selections[parent_key] != required_parent_value:
                return {
                    'valid': False,
                    'price': None,
                    'missing_dimensions': [],
                    'invalid_selections': [{
                        'dimension_key': dim.dimension_key,
                        'option_key': selections[dim.dimension_key],
                        'reason': (
                            f'维度 {dim.dimension_key} 仅在 {parent_key}={required_parent_value} '
                            f'时可选，但当前 {parent_key}={selections[parent_key]}'
                        ),
                    }],
                    'breakdown': {},
                    'reason': '级联约束不满足',
                }

        # 3. 模式 A：查映射表
        if product.pricing_mode == 'MATRIX':
            sig = ProductPriceMatrix.build_signature(selections)
            row = ProductPriceMatrix.objects.filter(
                product=product, config_signature=sig
            ).first()
            if not row:
                return {
                    'valid': False,
                    'price': None,
                    'missing_dimensions': [],
                    'invalid_selections': [],
                    'breakdown': selections,
                    'reason': '该配置组合无对应价格',
                }
            return {
                'valid': True,
                'price': row.price,
                'missing_dimensions': [],
                'invalid_selections': [],
                'breakdown': selections,
                'reason': '',
            }

        # 4. 模式 B：基准价 + 加价
        base = product.base_price or Decimal('0')
        price = base
        deltas = []
        for k, v in selections.items():
            rule = ProductPriceRule.objects.filter(
                product=product, dimension_key=k, option_key=v
            ).first()
            if rule:
                price += rule.price_delta
                deltas.append({'dimension': k, 'option': v, 'delta': str(rule.price_delta)})
        return {
            'valid': True,
            'price': price,
            'missing_dimensions': [],
            'invalid_selections': [],
            'breakdown': {'base_price': str(base), 'deltas': deltas},
            'reason': '',
        }


# ─── 产品配置 Excel 导入服务 ──────────────────────────────────────────────────

class ConfigExcelService:
    """产品配置 Excel 导入（OPT-6）"""

    @staticmethod
    def generate_template() -> bytes:
        """生成标准配置 Excel 模板"""
        wb = openpyxl.Workbook()

        # Sheet 1: dimensions
        ws1 = wb.active
        ws1.title = 'dimensions'
        ws1.append(['dimension_key', 'dimension_label', 'options', 'parent_dimension', 'is_required', 'sort_order'])
        ws1.append(['frame_color_back', '框架颜色（背框）', 'P1,P2,P3', '', 'TRUE', '1'])
        ws1.append(['backrest_material', '靠背材质', '网,布,皮', '', 'TRUE', '2'])
        ws1.append(['backrest_series', '靠背材质系列', 'AIR,3D knit,intermix', 'backrest_material=网', 'FALSE', '3'])
        ws1.append(['armrest', '扶手', '2D,3D,4D', '', 'TRUE', '4'])

        # Sheet 2: pricing_mode
        ws2 = wb.create_sheet('pricing_mode')
        ws2.append(['mode', 'base_price'])
        ws2.append(['MATRIX', ''])

        # Sheet 3a: matrix
        ws3 = wb.create_sheet('matrix')
        ws3.append(['frame_color_back', 'backrest_material', 'backrest_series', 'armrest', 'price'])
        ws3.append(['P1', '网', 'AIR', '3D', '2580'])

        # Sheet 3b: rules
        ws4 = wb.create_sheet('rules')
        ws4.append(['dimension_key', 'option_key', 'price_delta'])
        ws4.append(['frame_color_back', 'P1', '0'])
        ws4.append(['frame_color_back', 'P2', '50'])
        ws4.append(['backrest_material', '皮', '800'])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    @staticmethod
    def parse_excel(product: Product, file) -> dict:
        """
        解析上传的配置 Excel，返回预览结果。
        返回: {
            'pricing_mode': 'MATRIX'|'RULE',
            'base_price': Decimal|None,
            'dimensions': [...],
            'price_entries': [...],
            'errors': [...],
            'success_count': int,
            'failed_count': int,
        }
        """
        wb = openpyxl.load_workbook(file, read_only=True)
        errors = []
        result = {
            'pricing_mode': 'MATRIX',
            'base_price': None,
            'dimensions': [],
            'price_entries': [],
            'errors': errors,
            'success_count': 0,
            'failed_count': 0,
        }

        # Parse pricing_mode sheet
        if 'pricing_mode' in wb.sheetnames:
            ws = wb['pricing_mode']
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            if rows and rows[0]:
                mode = str(rows[0][0] or 'MATRIX').strip().upper()
                if mode in ('MATRIX', 'RULE'):
                    result['pricing_mode'] = mode
                base = rows[0][1] if len(rows[0]) > 1 else None
                if base:
                    try:
                        result['base_price'] = Decimal(str(base))
                    except Exception:
                        errors.append('pricing_mode sheet: base_price 格式错误')

        # Parse dimensions sheet
        if 'dimensions' in wb.sheetnames:
            ws = wb['dimensions']
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if len(rows) > 1:
                headers = [str(h or '').strip() for h in rows[0]]
                for i, row in enumerate(rows[1:], start=2):
                    data = dict(zip(headers, row))
                    dim_key = str(data.get('dimension_key', '') or '').strip()
                    if not dim_key:
                        errors.append(f'dimensions sheet 第{i}行: dimension_key 为空')
                        continue
                    options_raw = str(data.get('options', '') or '').strip()
                    options = [{'key': o.strip(), 'label': o.strip()} for o in options_raw.split(',') if o.strip()]
                    result['dimensions'].append({
                        'dimension_key': dim_key,
                        'dimension_label': str(data.get('dimension_label', '') or dim_key).strip(),
                        'options': options,
                        'parent_dimension': str(data.get('parent_dimension', '') or '').strip(),
                        'is_required': str(data.get('is_required', 'TRUE')).strip().upper() == 'TRUE',
                        'sort_order': int(data.get('sort_order', 0) or 0),
                    })

        # Parse matrix or rules sheet
        if result['pricing_mode'] == 'MATRIX' and 'matrix' in wb.sheetnames:
            ws = wb['matrix']
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if len(rows) > 1:
                headers = [str(h or '').strip() for h in rows[0]]
                price_col = headers.index('price') if 'price' in headers else -1
                dim_headers = [h for h in headers if h != 'price']
                for i, row in enumerate(rows[1:], start=2):
                    data = dict(zip(headers, row))
                    try:
                        price = Decimal(str(data.get('price', 0)))
                    except Exception:
                        errors.append(f'matrix sheet 第{i}行: price 格式错误')
                        result['failed_count'] += 1
                        continue
                    config = {h: str(data.get(h, '') or '').strip() for h in dim_headers}
                    result['price_entries'].append({'config': config, 'price': price})
                    result['success_count'] += 1

        elif result['pricing_mode'] == 'RULE' and 'rules' in wb.sheetnames:
            ws = wb['rules']
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if len(rows) > 1:
                for i, row in enumerate(rows[1:], start=2):
                    if len(row) < 3:
                        errors.append(f'rules sheet 第{i}行: 列数不足')
                        result['failed_count'] += 1
                        continue
                    dim_key = str(row[0] or '').strip()
                    opt_key = str(row[1] or '').strip()
                    try:
                        delta = Decimal(str(row[2]))
                    except Exception:
                        errors.append(f'rules sheet 第{i}行: price_delta 格式错误')
                        result['failed_count'] += 1
                        continue
                    result['price_entries'].append({
                        'dimension_key': dim_key,
                        'option_key': opt_key,
                        'price_delta': delta,
                    })
                    result['success_count'] += 1

        return result

    @staticmethod
    @transaction.atomic
    def execute_import(product: Product, parsed: dict):
        """确认导入：先清空后写入"""
        # 更新产品定价模式
        product.pricing_mode = parsed['pricing_mode']
        if parsed['base_price'] is not None:
            product.base_price = parsed['base_price']
        product.save(update_fields=['pricing_mode', 'base_price'])

        # 清空旧数据
        ProductConfigDimension.objects.filter(product=product).delete()
        ProductPriceMatrix.objects.filter(product=product).delete()
        ProductPriceRule.objects.filter(product=product).delete()

        # 写入维度
        for dim in parsed['dimensions']:
            ProductConfigDimension.objects.create(
                product=product,
                dimension_key=dim['dimension_key'],
                dimension_label=dim['dimension_label'],
                options=dim['options'],
                parent_dimension=dim['parent_dimension'],
                is_required=dim['is_required'],
                sort_order=dim['sort_order'],
            )

        # 写入价格数据
        if parsed['pricing_mode'] == 'MATRIX':
            for entry in parsed['price_entries']:
                sig = ProductPriceMatrix.build_signature(entry['config'])
                ProductPriceMatrix.objects.create(
                    product=product,
                    config_signature=sig,
                    config_attributes=entry['config'],
                    price=entry['price'],
                )
        else:  # RULE
            for i, entry in enumerate(parsed['price_entries']):
                ProductPriceRule.objects.create(
                    product=product,
                    dimension_key=entry['dimension_key'],
                    option_key=entry['option_key'],
                    price_delta=entry['price_delta'],
                    sort_order=i,
                )

        # 更新 min_price（取最低价格）
        if parsed['pricing_mode'] == 'MATRIX' and parsed['price_entries']:
            min_p = min(e['price'] for e in parsed['price_entries'])
            product.min_price = min_p
            product.save(update_fields=['min_price'])


# ─── 类别选项服务 ─────────────────────────────────────────────────────────────

class CategoryOptionsService:
    """一级 + 二级类别枚举（6 大类，中文标签）"""

    # 一级类别中文标签
    L1_LABELS = {
        'SEATING': '坐具类',
        'DESKS_WORKSTATIONS': '工位办公桌类',
        'TABLE': '桌台类',
        'STORAGE': '收纳储物类',
        'ACCESSORIES': '配套附件类',
        'EDUCATION': '教育家具类',
    }

    # 一级 → 二级映射（key 稳定英文，label 中文）
    L1_L2_MAP = {
        'SEATING': [
            ('OFFICE_CHAIR', '办公椅'),
            ('GUEST_CHAIR', '访客椅'),
            ('CONFERENCE_CHAIR', '会议椅'),
            ('STOOL', '凳子'),
            ('LOUNGE_SEATING', '休闲坐具'),
            ('VISITOR_SEATING', '接待椅'),
            ('OPERATOR_SEATING', '职员椅'),
        ],
        'DESKS_WORKSTATIONS': [
            ('DESK', '办公桌'),
            ('HEIGHT_ADJUSTABLE_DESK', '升降桌'),
            ('BENCHING', '屏风工位'),
            ('PRIVATE_OFFICE', '独立办公室办公桌'),
        ],
        'TABLE': [
            ('CONFERENCE_TABLE', '会议/协作桌'),
            ('OCCASIONAL_TABLE', '休闲桌'),
            ('OUTDOOR_TABLE', '户外桌及遮阳设施'),
        ],
        'STORAGE': [
            ('WORKSTATION_STORAGE', '工位收纳'),
            ('LOCKER', '储物柜'),
            ('CABINET_CREDENZA', '文件柜/矮柜'),
            ('BOOKCASE_SHELVING', '书柜/置物架'),
            ('CART', '移动推车'),
        ],
        'ACCESSORIES': [
            ('MODULAR_WALL', '模块化隔断及隔音材料'),
            ('POD', '独立 Pod 单元'),
            ('FREESTANDING_SCREEN', '独立屏风'),
            ('SPACE_DIVISION', '建筑/空间分割件'),
            ('MONITOR_ARM', '显示器支架及配件'),
            ('CABLE_MANAGEMENT', '电源/线缆管理'),
            ('LIGHTING', '照明设备'),
            ('ACC_TABLE', '附属桌台配件'),
        ],
        'EDUCATION': [
            ('CLASSROOM_CHAIR', '教室椅'),
            ('EDUCATION_LOUNGE', '教育休闲家具'),
            ('EDU_SEATING', '教育类坐具'),
            ('CLASSROOM_STORAGE', '教室收纳'),
            ('EDU_DESK', '教育类工位桌'),
            ('EDU_ACCESSORY', '教育类附件'),
        ],
    }

    @classmethod
    def get_options(cls) -> dict:
        l1_options = [{'value': k, 'label': cls.L1_LABELS.get(k, k)} for k in cls.L1_L2_MAP.keys()]
        l2_options = {}
        for l1, items in cls.L1_L2_MAP.items():
            l2_options[l1] = [{'value': v, 'label': label} for v, label in items]
        return {'category_l1': l1_options, 'category_l2': l2_options}


# ─── 批量产品导入（长格式，多产品单表） ──────────────────────────────────────

class BatchProductImportService:
    """
    批量产品导入 —— 采用客户长格式（竖排），支持一次导入多个产品。

    模板两个 Sheet：
      「产品」   : 产品类别 | 编号 | 一级分类 | 二级分类 | 品牌 | 产地 | 货期 | 形状 | 定价模式 | 描述
      「配置参数」: 产品类别 | 配置大类 | 部件/材质 | 规格/代码 | 价格 | 是否默认配置

    规则：
      - 「产品」每行 → 幂等创建/更新 Product（按编号优先，否则按名称匹配 name=产品类别）。
      - 「配置参数」按 (产品类别, 配置大类) 分组 → 每个配置大类 = 一个 ProductConfigDimension，
        行 = 选项（规格/代码 作 key，部件/材质+规格/代码 作 label）。
      - "配置款式" 类的行 → ProductConfigPreset（code=规格/代码，label=部件/材质）。
      - 是否默认配置=TRUE → 标记默认预设 / 默认选项。
      - 价格列可空：填了价则写入 ProductPriceRule（dimension_key, option_key, price_delta=价格）作为
        结构化留存；最终整机查表定价以后续确认的价格表为准。
    """

    ORIGIN_MAP = {'进口': 'IMPORT', '国产': 'DOMESTIC', 'IMPORT': 'IMPORT', 'DOMESTIC': 'DOMESTIC'}
    PRODUCT_SHEET = '产品'
    CONFIG_SHEET = '配置参数'
    STYLE_DIMENSION_KEYS = ('配置款式', '款式')  # 视为预设款式的配置大类名

    @staticmethod
    def _truthy(v) -> bool:
        return str(v or '').strip().upper() in ('TRUE', '1', '是', 'Y', 'YES', '默认')

    @classmethod
    def generate_template(cls) -> bytes:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = cls.PRODUCT_SHEET
        ws1.append(['产品类别', '编号', '一级分类', '二级分类', '品牌', '产地', '货期', '形状', '定价模式', '描述'])
        ws1.append(['Leader 升降经理桌', 'ZK-LEADER-01', 'DESKS_WORKSTATIONS', 'HEIGHT_ADJUSTABLE_DESK',
                    'ZIKOO', '国产', 'WITHIN_45D', '方形', 'MATRIX', '示例产品'])

        ws2 = wb.create_sheet(cls.CONFIG_SHEET)
        ws2.append(['产品类别', '配置大类', '部件/材质', '规格/代码', '价格', '是否默认配置'])
        ws2.append(['Leader 升降经理桌', '桌板材质', '科技木皮', 'VT-01 科技浅橡木', '', ''])
        ws2.append(['Leader 升降经理桌', '产品规格', '规格21', 'W2100*D1800*H650/1250', '', 'TRUE'])
        ws2.append(['Leader 升降经理桌', '配置款式', 'WW', '白色钢脚+白色边框+木皮台面', '', 'TRUE'])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    @classmethod
    def parse(cls, file) -> dict:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        errors = []
        products = {}  # name -> meta dict

        # 解析「产品」sheet
        if cls.PRODUCT_SHEET not in wb.sheetnames:
            errors.append(f'缺少「{cls.PRODUCT_SHEET}」sheet')
        else:
            ws = wb[cls.PRODUCT_SHEET]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if len(rows) > 1:
                headers = [str(h or '').strip() for h in rows[0]]
                for i, row in enumerate(rows[1:], start=2):
                    data = dict(zip(headers, row))
                    name = str(data.get('产品类别', '') or '').strip()
                    if not name:
                        continue
                    products[name] = {
                        'name': name,
                        'code': str(data.get('编号', '') or '').strip() or None,
                        'category_l1': str(data.get('一级分类', '') or '').strip() or 'SEATING',
                        'category_l2': str(data.get('二级分类', '') or '').strip(),
                        'brand': str(data.get('品牌', '') or '').strip(),
                        'origin': cls.ORIGIN_MAP.get(str(data.get('产地', '') or '').strip(), 'DOMESTIC'),
                        'lead_time': str(data.get('货期', '') or '').strip(),
                        'shape': str(data.get('形状', '') or '').strip(),
                        'pricing_mode': (str(data.get('定价模式', '') or 'MATRIX').strip().upper()
                                         if str(data.get('定价模式', '') or '').strip().upper() in ('MATRIX', 'RULE')
                                         else 'MATRIX'),
                        'description': str(data.get('描述', '') or '').strip(),
                        'dimensions': {},   # 配置大类 -> [ {key,label,price,is_default}, ... ]
                        'presets': [],      # [ {code,label,is_default} ]
                    }

        # 解析「配置参数」sheet
        if cls.CONFIG_SHEET not in wb.sheetnames:
            errors.append(f'缺少「{cls.CONFIG_SHEET}」sheet')
        else:
            ws = wb[cls.CONFIG_SHEET]
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if len(rows) > 1:
                headers = [str(h or '').strip() for h in rows[0]]
                for i, row in enumerate(rows[1:], start=2):
                    data = dict(zip(headers, row))
                    pname = str(data.get('产品类别', '') or '').strip()
                    big = str(data.get('配置大类', '') or '').strip()
                    part = str(data.get('部件/材质', '') or '').strip()
                    code = str(data.get('规格/代码', '') or '').strip()
                    price_raw = data.get('价格')
                    is_default = cls._truthy(data.get('是否默认配置'))
                    if not pname or not big:
                        continue
                    if pname not in products:
                        # 「产品」sheet 未声明的产品，自动补一个占位
                        products[pname] = {
                            'name': pname, 'code': None, 'category_l1': 'SEATING', 'category_l2': '',
                            'brand': '', 'origin': 'DOMESTIC', 'lead_time': '', 'shape': '',
                            'pricing_mode': 'MATRIX', 'description': '',
                            'dimensions': {}, 'presets': [],
                        }
                    p = products[pname]
                    price = None
                    if price_raw not in (None, ''):
                        try:
                            price = Decimal(str(price_raw))
                        except Exception:
                            errors.append(f'{cls.CONFIG_SHEET} 第{i}行: 价格格式错误({price_raw})')
                    opt_key = code or part
                    opt_label = (f'{code} {part}'.strip() if code and part else (code or part))
                    if big in cls.STYLE_DIMENSION_KEYS:
                        p['presets'].append({'code': opt_key, 'label': part or code, 'is_default': is_default})
                    else:
                        p['dimensions'].setdefault(big, [])
                        p['dimensions'][big].append({
                            'key': opt_key, 'label': opt_label, 'price': price, 'is_default': is_default,
                        })

        warnings = []
        if any(
            option.get('price') is not None
            for product in products.values()
            for options in product['dimensions'].values()
            for option in options
        ):
            warnings.append('旧版长格式的“价格”是单行值，无法代表完整组合最终价，预览时将忽略')
        if any(product['presets'] for product in products.values()):
            warnings.append('旧版“配置款式”未包含完整 selections，不会直接生成默认预设')
        summary = {
            'format': 'legacy',
            'product_count': len(products),
            'products': [
                {
                    'name': p['name'], 'code': p['code'],
                    'dimension_count': len(p['dimensions']),
                    'option_count': sum(len(v) for v in p['dimensions'].values()),
                    'preset_count': 0,
                    'price_count': 0,
                }
                for p in products.values()
            ],
            'errors': errors,
            'warnings': warnings,
        }
        return {'format': 'legacy', 'products': products, 'summary': summary,
                'errors': errors, 'warnings': warnings}

    @classmethod
    @transaction.atomic
    def execute_import(cls, parsed: dict, user):
        from .models import Brand as _Brand
        created, updated = 0, 0
        for meta in parsed['products'].values():
            # 品牌
            brand = None
            if meta['brand']:
                brand, _ = _Brand.objects.get_or_create(name=meta['brand'])
            # 产品（编号优先匹配）
            product = None
            if meta['code']:
                product = Product.objects.filter(code=meta['code']).first()
            if not product:
                product = Product.objects.filter(name=meta['name']).first()
            fields = dict(
                name=meta['name'], code=meta['code'],
                category_l1=meta['category_l1'], category_l2=meta['category_l2'],
                brand=brand, origin=meta['origin'], lead_time=meta['lead_time'],
                shape=meta['shape'], pricing_mode=meta['pricing_mode'],
                description=meta['description'],
            )
            if product:
                for k, v in fields.items():
                    setattr(product, k, v)
                product.save()
                updated += 1
            else:
                product = Product.objects.create(created_by=user, **fields)
                created += 1

            # 重建配置；旧长格式仅作为配置选项兼容导入，不再把“行价格”误当最终组合价。
            ProductConfigDimension.objects.filter(product=product).delete()
            ProductPriceMatrix.objects.filter(product=product).delete()
            ProductPriceRule.objects.filter(product=product).delete()
            ProductConfigPreset.objects.filter(product=product).delete()

            complete_default = {}
            for order, (big, opts) in enumerate(meta['dimensions'].items()):
                ProductConfigDimension.objects.create(
                    product=product, dimension_key=big, dimension_label=big,
                    options=[{'key': o['key'], 'label': o['label']} for o in opts],
                    is_required=True, sort_order=order,
                )
                defaults = [option for option in opts if option['is_default']]
                if len(defaults) == 1:
                    complete_default[big] = defaults[0]['key']
            # 只有每个必填维度都明确给出一个默认项时，才生成真实完整默认预设。
            if meta['dimensions'] and len(complete_default) == len(meta['dimensions']):
                ProductConfigPreset.objects.create(
                    product=product, code='LEGACY-DEFAULT', label='默认配置',
                    selections=ProductPriceMatrix.normalize_selections(complete_default),
                    is_default=True, sort_order=0,
                )
            product.pricing_mode = 'MATRIX'
            product.min_price = None
            product.save(update_fields=['pricing_mode', 'min_price'])
        return {
            'created': created,
            'updated': updated,
            'warnings': ['旧版长格式仅导入配置选项；行价格和不完整款式不会写入最终组合价格或默认预设'],
        }


# ─── 产品配置导出 ─────────────────────────────────────────────────────────────

class ConfigExportService:
    """导出单个产品当前配置数据为 Excel（PM-6）"""

    @staticmethod
    def export(product: Product) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'dimensions'
        ws.append(['dimension_key', 'dimension_label', 'options', 'parent_dimension', 'is_required', 'sort_order'])
        for d in product.config_dimensions.all():
            opts = ','.join(
                (o.get('key', '') if isinstance(o, dict) else str(o)) for o in (d.options or [])
            )
            ws.append([d.dimension_key, d.dimension_label, opts, d.parent_dimension,
                       'TRUE' if d.is_required else 'FALSE', d.sort_order])

        ws2 = wb.create_sheet('pricing')
        if product.pricing_mode == 'MATRIX':
            ws2.append(['config_signature', 'config_attributes', 'price'])
            for m in product.price_matrix.all():
                ws2.append([m.config_signature, json.dumps(m.config_attributes, ensure_ascii=False), str(m.price)])
        else:
            ws2.append(['dimension_key', 'option_key', 'price_delta'])
            for r in product.price_rules.all():
                ws2.append([r.dimension_key, r.option_key, str(r.price_delta)])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


# ─── 客户横向工作簿导入 ──────────────────────────────────────────────────────

class CustomerWorkbookImportService:
    """解析“每个产品一个 Sheet、每列一个维度”的客户自助配置模板。"""

    PRODUCT_INFO_SHEET = '产品信息'
    PRICE_SHEET = '组合价格'
    EMPTY_VALUES = {'', 'N/A', 'NA', '不适用', '-'}
    SECONDARY_MARKERS = ('二级', '2级')

    @staticmethod
    def _text(value) -> str:
        text = str(value or '').replace('\n', ' ').replace('\r', ' ').strip()
        text = ' '.join(text.split())
        return text.replace('（', '(').replace('）', ')')

    @classmethod
    def _is_empty(cls, value) -> bool:
        return cls._text(value).upper() in cls.EMPTY_VALUES

    @classmethod
    def _key(cls, label: str, used: set, prefix: str = '') -> str:
        raw = cls._text(label).lower()
        for ch in ('/', '\\', ' ', '-', '(', ')', '：', ':', ',', '，'):
            raw = raw.replace(ch, '_')
        raw = '_'.join(part for part in raw.split('_') if part)
        base = f'{prefix}_{raw}'.strip('_')[:90] or 'dimension'
        key = base
        suffix = 2
        while key in used:
            key = f'{base}_{suffix}'[:100]
            suffix += 1
        used.add(key)
        return key

    @classmethod
    def _default_name(cls, sheet_name: str) -> str:
        name = cls._text(sheet_name)
        if '(' in name and name.endswith(')'):
            inside = name.rsplit('(', 1)[1][:-1].strip()
            if inside:
                return inside
        for prefix in ('椅类模板', '桌类模板', '沙发', '休闲椅'):
            if name.startswith(prefix):
                name = name[len(prefix):].strip()
        return name or cls._text(sheet_name)

    @staticmethod
    def _infer_category(sheet_name: str) -> tuple[str, str]:
        if '桌' in sheet_name:
            return 'DESKS_WORKSTATIONS', 'FIXED_DESK'
        if '沙发' in sheet_name:
            return 'SEATING', 'LOUNGE_CHAIR'
        return 'SEATING', 'TASK_CHAIR'

    @classmethod
    def generate_template(cls) -> bytes:
        wb = openpyxl.Workbook()
        info = wb.active
        info.title = cls.PRODUCT_INFO_SHEET
        info.append(['配置Sheet', '产品编号', '产品名称', '一级分类', '二级分类', '品牌',
                     '产地', '货期', '形状', '描述'])
        info.append(['Think配置', 'THINK-001', 'Think办公椅', 'SEATING', 'TASK_CHAIR',
                     'Steelcase', '进口', 'WITHIN_45D', '', '示例产品'])

        options = wb.create_sheet('Think配置')
        options.append(['背框颜色', '坐垫饰面', '布（二级）', '皮（二级）', '扶手配置', '头枕'])
        options.append(['黑', '布', 'P1(OTTO)', 'A1(Leather)', '2D', '无'])
        options.append(['白', '皮', 'P2(Buzz2)', '', '4D', '有'])

        prices = wb.create_sheet(cls.PRICE_SHEET)
        prices.append(['配置Sheet', '组合编号', '组合名称', '是否默认', '最终价格',
                       '背框颜色', '坐垫饰面', '坐垫饰面_布', '坐垫饰面_皮', '扶手配置', '头枕'])
        prices.append(['Think配置', 'THINK-STD', '标准配置', '是', 3580,
                       '黑', '布', 'P1(OTTO)', '', '2D', '无'])
        prices.append(['Think配置', 'THINK-LEATHER', '皮面配置', '否', 4680,
                       '黑', '皮', '', 'A1(Leather)', '4D', '有'])

        notes = wb.create_sheet('填写说明')
        notes.append(['说明'])
        notes.append(['每个产品配置 Sheet：第一行是配置维度，下面逐列填写可选项。'])
        notes.append(['“布（二级）/皮（二级）”自动依赖前一个饰面/材质列；N/A 和空白不会导入。'])
        notes.append(['组合价格：每一行必须是一套完整有效配置；是否默认每个产品最多一行“是”。'])
        notes.append(['新增未知列无需改程序，上传预览会按列自动生成维度并显示解析结果。'])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    @classmethod
    def parse(cls, file) -> dict:
        # 使用普通模式以读取隐藏列和合并单元格信息；客户模板体量很小。
        wb = openpyxl.load_workbook(file, read_only=False, data_only=True)
        errors, warnings = [], []
        info_by_sheet = {}
        if cls.PRODUCT_INFO_SHEET in wb.sheetnames:
            rows = list(wb[cls.PRODUCT_INFO_SHEET].iter_rows(values_only=True))
            if rows:
                headers = [cls._text(v) for v in rows[0]]
                for row_no, row in enumerate(rows[1:], start=2):
                    data = dict(zip(headers, row))
                    sheet = cls._text(data.get('配置Sheet'))
                    if not sheet:
                        continue
                    if sheet not in wb.sheetnames:
                        errors.append(f'{cls.PRODUCT_INFO_SHEET}!A{row_no}: 配置Sheet“{sheet}”不存在')
                        continue
                    if sheet in info_by_sheet:
                        errors.append(f'{cls.PRODUCT_INFO_SHEET}!A{row_no}: 配置Sheet“{sheet}”重复')
                        continue
                    info_by_sheet[sheet] = data
        else:
            warnings.append('缺少“产品信息”Sheet：产品名称和分类将根据配置 Sheet 名称自动推断')

        ignored = {cls.PRODUCT_INFO_SHEET, cls.PRICE_SHEET, '填写说明'}
        config_sheets = [name for name in wb.sheetnames if name not in ignored and not name.endswith('组合价格')]
        products = {}
        conditional_children = {
            '坐垫饰面': {'布', '皮'},
            '背饰面': {'布', '皮', '网'},
            '软包类型': {'布', '皮'},
            '屏风材质': {'布', 'PET', '钢'},
        }
        for sheet_name in config_sheets:
            ws = wb[sheet_name]
            headers = [cls._text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
            used, dimensions, header_map = set(), {}, {}
            parent_key = ''
            parent_label = ''
            for col, header in enumerate(headers, start=1):
                if not header:
                    continue
                column_letter = openpyxl.utils.get_column_letter(col)
                if ws.column_dimensions[column_letter].hidden:
                    warnings.append(f'{sheet_name}!{column_letter}:{column_letter}: 隐藏列“{header}”已忽略')
                    continue
                options = []
                seen = set()
                for row in range(2, ws.max_row + 1):
                    value = cls._text(ws.cell(row, col).value)
                    if cls._is_empty(value) or value in seen:
                        continue
                    seen.add(value)
                    options.append({'key': value, 'label': value})
                if not options:
                    continue
                plain_header = header.split('(')[0].strip()
                explicit_secondary = any(marker in header for marker in cls.SECONDARY_MARKERS)
                contextual_secondary = (
                    bool(parent_key)
                    and parent_label in conditional_children
                    and plain_header in conditional_children[parent_label]
                )
                secondary = bool(parent_key) and (explicit_secondary or contextual_secondary)
                if secondary:
                    parent_value = plain_header
                    key = cls._key(parent_value, used, parent_key)
                    parent = f'{parent_key}={parent_value}'
                    label = f'{parent_label}-{parent_value}'
                else:
                    key = cls._key(header, used)
                    parent = ''
                    label = header
                    parent_key = key
                    parent_label = label
                dimensions[key] = {
                    'key': key, 'label': label, 'options': options,
                    'parent_dimension': parent, 'is_required': True,
                    'sort_order': len(dimensions), 'source_header': header,
                }
                # 重复的“布（二级）”等裸标题存在歧义时，不允许价格表用裸标题误匹配；
                # 始终提供“父维度_子值”限定别名。
                if header in header_map and header_map[header] != key:
                    header_map[header] = None
                else:
                    header_map[header] = key
                header_map[key] = key
                if secondary:
                    header_map[f'{parent_label}_{parent_value}'] = key
                    header_map[f'{parent_label}-{parent_value}'] = key

            meta = info_by_sheet.get(sheet_name, {})
            category_l1, category_l2 = cls._infer_category(sheet_name)
            name = cls._text(meta.get('产品名称')) or cls._default_name(sheet_name)
            code = cls._text(meta.get('产品编号')) or f'IMP-{hashlib.sha1(sheet_name.encode()).hexdigest()[:8].upper()}'
            products[sheet_name] = {
                'sheet_name': sheet_name, 'name': name, 'code': code,
                'category_l1': cls._text(meta.get('一级分类')) or category_l1,
                'category_l2': cls._text(meta.get('二级分类')) or category_l2,
                'brand': cls._text(meta.get('品牌')),
                'origin': BatchProductImportService.ORIGIN_MAP.get(cls._text(meta.get('产地')), 'DOMESTIC'),
                'lead_time': cls._text(meta.get('货期')),
                'shape': cls._text(meta.get('形状')),
                'pricing_mode': 'MATRIX', 'description': cls._text(meta.get('描述')),
                'dimensions': dimensions, 'header_map': header_map,
                'presets': [], 'price_matrix': [],
                '_preset_codes': set(), '_price_signatures': set(),
            }
            if not dimensions:
                warnings.append(f'{sheet_name}: 未发现有选项的配置列')

        seen_codes, seen_names = {}, {}
        for product in products.values():
            code_key = product['code'].casefold()
            name_key = product['name'].casefold()
            if code_key in seen_codes:
                errors.append(
                    f'{product["sheet_name"]}: 产品编号“{product["code"]}”与配置 Sheet“{seen_codes[code_key]}”重复')
            else:
                seen_codes[code_key] = product['sheet_name']
            if name_key in seen_names:
                errors.append(
                    f'{product["sheet_name"]}: 产品名称“{product["name"]}”与配置 Sheet“{seen_names[name_key]}”重复')
            else:
                seen_names[name_key] = product['sheet_name']
            if not Product.objects.filter(code=product['code']).exists():
                name_matches = Product.objects.filter(name=product['name'])
                if name_matches.count() > 1:
                    errors.append(
                        f'{product["sheet_name"]}: 产品名称“{product["name"]}”匹配到多条已有产品，请填写唯一产品编号')

        price_sheets = []
        if cls.PRICE_SHEET in wb.sheetnames:
            price_sheets.append((cls.PRICE_SHEET, None))
        price_sheets.extend((name, name[:-4]) for name in wb.sheetnames if name.endswith('组合价格') and name != cls.PRICE_SHEET)
        for price_sheet, fixed_product_sheet in price_sheets:
            ws = wb[price_sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [cls._text(v) for v in rows[0]]
            duplicate_headers = sorted({header for header in headers if header and headers.count(header) > 1})
            if duplicate_headers:
                errors.append(f'{price_sheet}: 表头重复：{"、".join(duplicate_headers)}')
                continue
            for row_no, row in enumerate(rows[1:], start=2):
                data = dict(zip(headers, row))
                target_sheet = fixed_product_sheet or cls._text(data.get('配置Sheet'))
                if not target_sheet:
                    continue
                product = products.get(target_sheet)
                if not product:
                    errors.append(f'{price_sheet} 第{row_no}行: 配置Sheet“{target_sheet}”不存在')
                    continue
                try:
                    price = Decimal(str(data.get('最终价格')))
                except Exception:
                    errors.append(f'{price_sheet} 第{row_no}行: 最终价格无效')
                    continue
                selections = {}
                row_invalid = False
                for header, value in data.items():
                    normalized = cls._text(value)
                    normalized_header = cls._text(header)
                    key = product['header_map'].get(normalized_header)
                    if (normalized_header in product['header_map'] and key is None
                            and not cls._is_empty(normalized)):
                        errors.append(
                            f'{price_sheet} 第{row_no}行: 配置列“{normalized_header}”有歧义，请使用带父维度的限定列名')
                        row_invalid = True
                        continue
                    if key and not cls._is_empty(normalized):
                        allowed = {o['key'] for o in product['dimensions'][key]['options']}
                        if normalized not in allowed:
                            errors.append(f'{price_sheet}!{openpyxl.utils.get_column_letter(headers.index(header) + 1)}{row_no}: {header} 的选项“{normalized}”不在选项库')
                            row_invalid = True
                        selections[key] = normalized
                selections = ProductPriceMatrix.normalize_selections(selections)
                if not selections:
                    errors.append(f'{price_sheet} 第{row_no}行: 未填写配置选项')
                    continue
                missing = []
                invisible = []
                for dim in sorted(product['dimensions'].values(), key=lambda item: item['sort_order']):
                    parent = dim['parent_dimension']
                    visible = True
                    if parent:
                        parent_key, _, parent_value = parent.partition('=')
                        visible = parent_key in selections and (
                            not parent_value or selections.get(parent_key) == parent_value
                        )
                    if visible and dim['is_required'] and dim['key'] not in selections:
                        missing.append(dim['label'])
                    if not visible and dim['key'] in selections:
                        invisible.append(dim['label'])
                if missing:
                    errors.append(f'{price_sheet} 第{row_no}行: 缺少当前条件下必填配置：{"、".join(missing)}')
                    row_invalid = True
                if invisible:
                    errors.append(f'{price_sheet} 第{row_no}行: 包含当前条件下不可见配置：{"、".join(invisible)}')
                    row_invalid = True
                if row_invalid:
                    continue
                code = cls._text(data.get('组合编号')) or f'COMBO-{row_no}'
                label = cls._text(data.get('组合名称')) or code
                signature = ProductPriceMatrix.build_signature(selections)
                if code.casefold() in product['_preset_codes']:
                    errors.append(f'{price_sheet} 第{row_no}行: 组合编号“{code}”重复')
                    continue
                if signature in product['_price_signatures']:
                    errors.append(f'{price_sheet} 第{row_no}行: 配置组合与前面的价格行重复')
                    continue
                product['_preset_codes'].add(code.casefold())
                product['_price_signatures'].add(signature)
                is_default = BatchProductImportService._truthy(data.get('是否默认'))
                product['price_matrix'].append({'config': selections, 'price': price})
                product['presets'].append({
                    'code': code, 'label': label, 'selections': selections,
                    'is_default': is_default,
                })

        for product in products.values():
            defaults = [p for p in product['presets'] if p['is_default']]
            if len(defaults) > 1:
                errors.append(f'{product["sheet_name"]}: 默认组合超过一条')
            if not product['price_matrix']:
                warnings.append(f'{product["sheet_name"]}: 暂无组合价格，只能导入配置选项，导入后不能加入报价单')
            if product['price_matrix'] and not defaults:
                warnings.append(f'{product["sheet_name"]}: 未设置默认组合，产品详情将默认进入自定义配置')

        summary_products = []
        for product in products.values():
            summary_products.append({
                'name': product['name'], 'code': product['code'], 'sheet_name': product['sheet_name'],
                'dimension_count': len(product['dimensions']),
                'option_count': sum(len(d['options']) for d in product['dimensions'].values()),
                'preset_count': len(product['presets']),
                'price_count': len(product['price_matrix']),
            })
        summary = {
            'format': 'horizontal', 'product_count': len(products),
            'products': summary_products, 'errors': errors, 'warnings': warnings,
        }
        return {'format': 'horizontal', 'products': products, 'summary': summary,
                'errors': errors, 'warnings': warnings}

    @classmethod
    @transaction.atomic
    def execute_import(cls, parsed: dict, user) -> dict:
        created = updated = 0
        for meta in parsed['products'].values():
            brand = None
            if meta['brand']:
                brand, _ = Brand.objects.get_or_create(name=meta['brand'])
            product = Product.objects.filter(code=meta['code']).first()
            if not product:
                name_matches = Product.objects.filter(name=meta['name'])
                if name_matches.count() > 1:
                    raise ValueError(f'产品名称“{meta["name"]}”匹配到多条已有产品，请填写唯一产品编号')
                product = name_matches.first()
            incoming_has_prices = bool(meta['price_matrix'])
            fields = {
                'name': meta['name'], 'code': meta['code'], 'category_l1': meta['category_l1'],
                'category_l2': meta['category_l2'], 'brand': brand, 'origin': meta['origin'],
                'lead_time': meta['lead_time'], 'shape': meta['shape'],
                'description': meta['description'],
            }
            if product is None or incoming_has_prices:
                fields['pricing_mode'] = 'MATRIX'
            if product:
                for key, value in fields.items():
                    setattr(product, key, value)
                product.save()
                updated += 1
            else:
                product = Product.objects.create(created_by=user, **fields)
                created += 1

            # 选项库始终按工作簿更新；只有工作簿包含完整组合价格时才替换价格和预设。
            # 这样客户原始“仅选项”文件不会清空线上已有可报价数据。
            product.config_dimensions.all().delete()
            if incoming_has_prices:
                product.price_matrix.all().delete()
                product.price_rules.all().delete()
                product.config_presets.all().delete()
            for dim in meta['dimensions'].values():
                ProductConfigDimension.objects.create(
                    product=product, dimension_key=dim['key'], dimension_label=dim['label'],
                    options=dim['options'], parent_dimension=dim['parent_dimension'],
                    is_required=dim['is_required'], sort_order=dim['sort_order'],
                )
            if incoming_has_prices:
                for entry in meta['price_matrix']:
                    config = ProductPriceMatrix.normalize_selections(entry['config'])
                    ProductPriceMatrix.objects.create(
                        product=product, config_signature=ProductPriceMatrix.build_signature(config),
                        config_attributes=config, price=entry['price'],
                    )
                for order, preset in enumerate(meta['presets']):
                    ProductConfigPreset.objects.create(
                        product=product, code=preset['code'], label=preset['label'],
                        selections=ProductPriceMatrix.normalize_selections(preset['selections']),
                        is_default=preset['is_default'], sort_order=order,
                    )
                prices = [entry['price'] for entry in meta['price_matrix']]
                product.min_price = min(prices)
                product.save(update_fields=['min_price'])
        return {'created': created, 'updated': updated, 'warnings': parsed.get('warnings', [])}


# ─── 事务化产品组合创建 ──────────────────────────────────────────────────────

class ProductCompositeService:
    @staticmethod
    @transaction.atomic
    def create(validated: dict, user, files=None) -> Product:
        product_data = validated['product']
        product = Product.objects.create(created_by=user, **product_data)
        dimensions = {}
        for order, dim in enumerate(validated.get('dimensions', [])):
            item = ProductConfigDimension.objects.create(
                product=product, sort_order=dim.get('sort_order', order), **{
                    key: value for key, value in dim.items() if key != 'sort_order'
                },
            )
            dimensions[item.dimension_key] = item

        matrix_prices = []
        for entry in validated.get('price_matrix', []):
            config = ProductPriceMatrix.normalize_selections(entry.get('config', {}))
            if not config:
                raise ValueError('价格组合必须包含配置选项')
            price = Decimal(str(entry.get('price')))
            ProductPriceMatrix.objects.create(
                product=product, config_signature=ProductPriceMatrix.build_signature(config),
                config_attributes=config, price=price,
            )
            matrix_prices.append(price)

        for order, preset in enumerate(validated.get('presets', [])):
            ProductConfigPreset.objects.create(
                product=product, code=preset['code'], label=preset.get('label', ''),
                selections=ProductPriceMatrix.normalize_selections(preset.get('selections', {})),
                is_default=preset.get('is_default', False),
                sort_order=preset.get('sort_order', order),
            )
        if matrix_prices:
            product.min_price = min(matrix_prices)
            product.save(update_fields=['min_price'])
        if files:
            ProductImageService.upload_images(product, list(files))
        return product


# ─── 安全配置维度变更与通用 Excel 导入（2026-08）────────────────────────────

class DimensionMutationService:
    """配置维度的引用分析、事务化修改与安全删除。"""

    @staticmethod
    def _normalize_options(options):
        normalized = []
        seen = set()
        for raw in options or []:
            if not isinstance(raw, dict):
                raise ValueError('选项必须包含 key 和 label')
            key = str(raw.get('key') or '').strip()
            label = str(raw.get('label') or key).strip()
            if not key:
                raise ValueError('选项键不能为空')
            if key in seen:
                raise ValueError(f'选项键“{key}”重复')
            seen.add(key)
            normalized.append({'key': key, 'label': label or key})
        if not normalized:
            raise ValueError('配置维度至少需要一个选项')
        return normalized

    @classmethod
    def impact(cls, dimension):
        from quotes.models import QuoteItem

        key = dimension.dimension_key
        matrices = [item for item in dimension.product.price_matrix.all()
                    if key in (item.config_attributes or {})]
        presets = [item for item in dimension.product.config_presets.all()
                   if key in (item.selections or {})]
        children = list(dimension.product.config_dimensions.filter(
            Q(parent_dimension=key) | Q(parent_dimension__startswith=f'{key}=')
        ))
        quote_items = [item for item in QuoteItem.objects.filter(product=dimension.product).only('config_attributes')
                       if key in (item.config_attributes or {})]
        rules = list(dimension.product.price_rules.filter(dimension_key=key))
        referenced_options = set()
        for item in matrices:
            referenced_options.add(str((item.config_attributes or {}).get(key, '')))
        for item in presets:
            referenced_options.add(str((item.selections or {}).get(key, '')))
        referenced_options.update(str(item.option_key) for item in rules)
        for child in children:
            _, _, value = child.parent_dimension.partition('=')
            if value:
                referenced_options.add(value)
        referenced_options.discard('')
        blocking = len(matrices) + len(presets) + len(children) + len(rules)
        return {
            'dimension_id': dimension.id,
            'dimension_key': key,
            'matrix_rows': len(matrices),
            'rules': len(rules),
            'presets': len(presets),
            'child_dimensions': len(children),
            'quote_items': len(quote_items),
            'can_delete': blocking == 0,
            'key_locked': blocking > 0,
            'locked_option_keys': sorted(referenced_options),
            'history_note': '历史报价保存配置和价格快照，维度变更不会改写历史报价。',
        }

    @classmethod
    @transaction.atomic
    def update(cls, product, dimension_id, data):
        dimension = ProductConfigDimension.objects.select_for_update().get(
            id=dimension_id, product=product)
        impact = cls.impact(dimension)
        old_key = dimension.dimension_key
        new_key = str(data.get('dimension_key', old_key)).strip()
        if not new_key:
            raise ValueError('维度键不能为空')
        if new_key != old_key and impact['key_locked']:
            raise ValueError('该维度已被定价、预设或下级维度引用，不能直接修改维度键')
        if ProductConfigDimension.objects.filter(product=product, dimension_key=new_key).exclude(id=dimension.id).exists():
            raise ValueError(f'维度键“{new_key}”已存在')

        options = cls._normalize_options(data.get('options', dimension.options))
        option_keys = {item['key'] for item in options}
        removed_locked = set(impact['locked_option_keys']) - option_keys
        if removed_locked:
            raise ValueError(f'选项 {"、".join(sorted(removed_locked))} 已被定价或级联引用，不能直接删除')

        parent = str(data.get('parent_dimension', dimension.parent_dimension) or '').strip()
        if parent:
            parent_key, _, parent_value = parent.partition('=')
            if parent_key == new_key:
                raise ValueError('维度不能依赖自身')
            parent_dim = ProductConfigDimension.objects.filter(
                product=product, dimension_key=parent_key).exclude(id=dimension.id).first()
            if not parent_dim:
                raise ValueError(f'父维度“{parent_key}”不存在')
            if parent_value and parent_value not in {str(o.get('key')) for o in parent_dim.options or []}:
                raise ValueError(f'父维度选项“{parent_value}”不存在')
            visited = {new_key}
            cursor = parent_dim
            while cursor and cursor.parent_dimension:
                ancestor_key = cursor.parent_dimension.partition('=')[0]
                if ancestor_key in visited:
                    raise ValueError('配置维度不能形成循环依赖')
                visited.add(ancestor_key)
                cursor = ProductConfigDimension.objects.filter(
                    product=product, dimension_key=ancestor_key).first()

        dimension.dimension_key = new_key
        dimension.dimension_label = str(data.get('dimension_label', dimension.dimension_label)).strip() or new_key
        dimension.options = options
        dimension.parent_dimension = parent
        dimension.is_required = bool(data.get('is_required', dimension.is_required))
        dimension.sort_order = int(data.get('sort_order', dimension.sort_order) or 0)
        dimension.save()
        return dimension

    @classmethod
    @transaction.atomic
    def delete(cls, product, dimension_id, force=False):
        dimension = ProductConfigDimension.objects.select_for_update().get(
            id=dimension_id, product=product)
        impact = cls.impact(dimension)
        if impact['child_dimensions']:
            raise ValueError('该维度仍有下级维度，请先修改或删除下级维度的级联关系')
        if not impact['can_delete'] and not force:
            raise ValueError('该维度仍被定价或默认配置引用，请查看影响后再确认强制删除')

        removed = {'matrix_rows': 0, 'rules': 0, 'presets': 0}
        if force:
            key = dimension.dimension_key
            matrix_ids = [item.id for item in product.price_matrix.all()
                          if key in (item.config_attributes or {})]
            preset_ids = [item.id for item in product.config_presets.all()
                          if key in (item.selections or {})]
            removed['matrix_rows'], _ = product.price_matrix.filter(id__in=matrix_ids).delete()
            removed['presets'], _ = product.config_presets.filter(id__in=preset_ids).delete()
            removed['rules'], _ = product.price_rules.filter(dimension_key=key).delete()
        dimension.delete()
        lowest = product.price_matrix.order_by('price').values_list('price', flat=True).first()
        product.min_price = lowest
        product.save(update_fields=['min_price'])
        return {'impact': impact, 'removed': removed}


class FlexibleConfigExcelService:
    """兼容标准、自制横向、自制纵向及组合价格表的安全配置导入。"""

    SHEET_ALIASES = {
        'dimensions': {'dimensions', '维度', '配置维度', '维度配置'},
        'pricing_mode': {'pricingmode', '定价模式', '价格模式'},
        'matrix': {'matrix', '组合价格', '价格矩阵', '配置价格'},
        'rules': {'rules', '价格规则', '加价规则'},
        'presets': {'presets', '默认配置', '配置预设', '预设配置'},
    }
    HEADER_ALIASES = {
        'dimension_key': {'dimensionkey', '维度键', '配置键'},
        'dimension_label': {'dimensionlabel', '维度名称', '配置名称', '配置维度', '维度'},
        'options': {'options', '选项', '配置选项', '选项列表'},
        'option': {'option', 'optionlabel', '选项', '选项名称', '配置选项'},
        'parent_dimension': {'parentdimension', '父维度', '级联', '级联条件'},
        'is_required': {'isrequired', 'required', '是否必填', '必填'},
        'sort_order': {'sortorder', '排序', '顺序'},
        'price': {'price', '最终价格', '组合价格', '价格', '售价'},
        'price_delta': {'pricedelta', '加价', '价格增量'},
        'option_key': {'optionkey', '选项键', '选项代码'},
        'code': {'code', '组合编号', '预设编号'},
        'label': {'label', '组合名称', '预设名称'},
        'is_default': {'isdefault', '是否默认', '默认'},
    }
    META_HEADERS = {'code', 'label', 'is_default'}

    @staticmethod
    def _text(value):
        return ' '.join(str(value or '').replace('\n', ' ').replace('\r', ' ').strip().split())

    @classmethod
    def _norm(cls, value):
        text = cls._text(value).casefold()
        for char in (' ', '_', '-', '/', '\\', '（', '）', '(', ')', '：', ':'):
            text = text.replace(char, '')
        return text

    @classmethod
    def _canonical_header(cls, value):
        normalized = cls._norm(value)
        for canonical, aliases in cls.HEADER_ALIASES.items():
            if normalized in {cls._norm(alias) for alias in aliases}:
                return canonical
        return ''

    @classmethod
    def _sheet_map(cls, workbook):
        result = {}
        for name in workbook.sheetnames:
            normalized = cls._norm(name)
            for canonical, aliases in cls.SHEET_ALIASES.items():
                if normalized in {cls._norm(alias) for alias in aliases}:
                    result[canonical] = name
        return result

    @classmethod
    def _dimension_key(cls, label, used):
        raw = cls._text(label).lower()
        safe = ''.join(char if char.isascii() and char.isalnum() else '_' for char in raw)
        safe = '_'.join(part for part in safe.split('_') if part)
        base = safe[:80] or f'dim_{hashlib.sha1(raw.encode()).hexdigest()[:10]}'
        key, index = base, 2
        while key in used:
            key = f'{base}_{index}'[:100]
            index += 1
        used.add(key)
        return key

    @staticmethod
    def _truthy(value):
        return str(value or '').strip().upper() in {'TRUE', '1', '是', 'Y', 'YES', '默认'}

    @classmethod
    def _split_options(cls, value):
        import re
        items, seen = [], set()
        for token in re.split(r'[,，;；\n]+', cls._text(value)):
            token = token.strip()
            if not token or token in seen:
                continue
            seen.add(token)
            if '|' in token:
                key, label = (part.strip() for part in token.split('|', 1))
            else:
                key = label = token
            if key:
                items.append({'key': key, 'label': label or key})
        return items

    @classmethod
    def _base_result(cls, workbook):
        available = []
        for name in workbook.sheetnames:
            ws = workbook[name]
            headers = [cls._text(cell.value) for cell in ws[1] if cls._text(cell.value)]
            preview = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
                preview.append([cls._text(value) for value in row])
            available.append({'name': name, 'headers': headers, 'preview': preview})
        return {
            'pricing_mode': 'MATRIX', 'base_price': None, 'dimensions': [],
            'price_entries': [], 'presets': [], 'errors': [], 'warnings': [],
            'success_count': 0, 'failed_count': 0, 'detected_format': '',
            'needs_mapping': False, 'available_sheets': available,
        }

    @classmethod
    def parse_excel(cls, product, file, mapping=None):
        mapping = mapping or {}
        workbook = openpyxl.load_workbook(file, read_only=False, data_only=True)
        result = cls._base_result(workbook)
        sheet_map = cls._sheet_map(workbook)
        if 'dimensions' in sheet_map and not mapping:
            cls._parse_standard(workbook, sheet_map, result)
        else:
            candidates = [item for item in result['available_sheets']
                          if item['headers'] and cls._norm(item['name']) not in {'填写说明', '说明'}]
            selected_name = mapping.get('sheet')
            if not selected_name and len(candidates) == 1:
                selected_name = candidates[0]['name']
            if not selected_name or selected_name not in workbook.sheetnames:
                result['needs_mapping'] = True
                result['detected_format'] = 'mapping_required'
                result['warnings'].append('检测到多个或无法确定的数据 Sheet，请选择 Sheet 和数据结构后重新解析')
                return cls._finish(product, result)
            cls._parse_custom(workbook[selected_name], result, mapping)
        return cls._finish(product, result)

    @classmethod
    def _parse_standard(cls, workbook, sheet_map, result):
        result['detected_format'] = 'standard'
        ws = workbook[sheet_map['dimensions']]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result['errors'].append(f'{ws.title}: 没有表头')
            return
        header_lookup = {cls._canonical_header(value): index for index, value in enumerate(rows[0])
                         if cls._canonical_header(value)}
        for required in ('dimension_key', 'dimension_label', 'options'):
            if required not in header_lookup:
                result['errors'].append(f'{ws.title}: 缺少必填表头 {required}')
        if result['errors']:
            return
        used = set()
        for row_no, row in enumerate(rows[1:], start=2):
            key = cls._text(row[header_lookup['dimension_key']] if len(row) > header_lookup['dimension_key'] else '')
            label = cls._text(row[header_lookup['dimension_label']] if len(row) > header_lookup['dimension_label'] else '')
            if not key and not label:
                continue
            if not key:
                result['errors'].append(f'{ws.title}!A{row_no}: 维度键为空')
                continue
            if key in used:
                result['errors'].append(f'{ws.title}!A{row_no}: 维度键“{key}”重复')
                continue
            used.add(key)
            options = cls._split_options(row[header_lookup['options']] if len(row) > header_lookup['options'] else '')
            if not options:
                result['errors'].append(f'{ws.title} 第{row_no}行: 选项为空')
                continue
            def value(name, default=''):
                index = header_lookup.get(name)
                return row[index] if index is not None and len(row) > index else default
            try:
                sort_order = int(value('sort_order', len(result['dimensions'])) or len(result['dimensions']))
            except (TypeError, ValueError):
                result['errors'].append(f'{ws.title} 第{row_no}行: 排序必须是整数')
                continue
            result['dimensions'].append({
                'dimension_key': key, 'dimension_label': label or key, 'options': options,
                'parent_dimension': cls._text(value('parent_dimension')),
                'is_required': cls._truthy(value('is_required', 'TRUE')),
                'sort_order': sort_order,
            })
        dimensions = {item['dimension_key']: item for item in result['dimensions']}
        if 'pricing_mode' in sheet_map:
            rows = list(workbook[sheet_map['pricing_mode']].iter_rows(min_row=2, values_only=True))
            if rows:
                mode = cls._text(rows[0][0]).upper()
                result['pricing_mode'] = mode if mode in {'MATRIX', 'RULE'} else 'MATRIX'
                if len(rows[0]) > 1 and rows[0][1] not in (None, ''):
                    try:
                        result['base_price'] = Decimal(str(rows[0][1]))
                    except Exception:
                        result['errors'].append(f'{sheet_map["pricing_mode"]}!B2: 基准价格式错误')
        target = sheet_map.get('matrix' if result['pricing_mode'] == 'MATRIX' else 'rules')
        if target:
            cls._parse_price_sheet(workbook[target], result, dimensions)
        if 'presets' in sheet_map:
            cls._parse_preset_sheet(workbook[sheet_map['presets']], result, dimensions)

    @classmethod
    def _parse_custom(cls, ws, result, mapping):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result['errors'].append(f'{ws.title}: 工作表为空')
            return
        headers = [cls._text(value) for value in rows[0]]
        canonical = {cls._canonical_header(header): header for header in headers if cls._canonical_header(header)}
        format_name = mapping.get('format')
        if not format_name:
            if ('dimension_label' in canonical and ('option' in canonical or 'options' in canonical)):
                format_name = 'vertical'
            elif 'price' in canonical:
                format_name = 'combination'
            else:
                format_name = 'horizontal'
        result['detected_format'] = format_name
        if format_name == 'vertical':
            dim_header = mapping.get('dimension_column') or canonical.get('dimension_label')
            option_header = mapping.get('option_column') or canonical.get('option') or canonical.get('options')
            if dim_header not in headers or option_header not in headers:
                result['needs_mapping'] = True
                result['errors'].append('纵向表需要指定“维度名称列”和“选项列”')
                return
            grouped, used = {}, set()
            for row_no, row in enumerate(rows[1:], start=2):
                data = dict(zip(headers, row))
                label = cls._text(data.get(dim_header))
                option_text = cls._text(data.get(option_header))
                if not label and not option_text:
                    continue
                if not label or not option_text:
                    result['errors'].append(f'{ws.title} 第{row_no}行: 维度名称或选项为空')
                    continue
                if label not in grouped:
                    grouped[label] = {
                        'dimension_key': cls._dimension_key(label, used), 'dimension_label': label,
                        'options': [], 'parent_dimension': '', 'is_required': True,
                        'sort_order': len(grouped),
                    }
                for option in cls._split_options(option_text):
                    if option['key'] not in {item['key'] for item in grouped[label]['options']}:
                        grouped[label]['options'].append(option)
                required_col = mapping.get('required_column') or canonical.get('is_required')
                parent_col = mapping.get('parent_column') or canonical.get('parent_dimension')
                if required_col in headers:
                    grouped[label]['is_required'] = cls._truthy(data.get(required_col))
                if parent_col in headers:
                    grouped[label]['parent_dimension'] = cls._text(data.get(parent_col))
            result['dimensions'] = list(grouped.values())
            return

        price_header = mapping.get('price_column') or canonical.get('price')
        metadata = {value for key, value in canonical.items() if key in cls.META_HEADERS}
        if price_header:
            metadata.add(price_header)
            format_name = result['detected_format'] = 'combination'
        dimension_headers = [header for header in headers if header and header not in metadata]
        if not dimension_headers:
            result['errors'].append(f'{ws.title}: 没有可识别的配置维度列')
            return
        used = set()
        dimensions = []
        for order, header in enumerate(dimension_headers):
            options, seen = [], set()
            for row in rows[1:]:
                data = dict(zip(headers, row))
                value = cls._text(data.get(header))
                if value and value.upper() not in {'N/A', 'NA', '-'} and value not in seen:
                    seen.add(value)
                    options.append({'key': value, 'label': value})
            if options:
                dimensions.append({
                    'dimension_key': cls._dimension_key(header, used),
                    'dimension_label': header, 'options': options,
                    'parent_dimension': '', 'is_required': True, 'sort_order': order,
                    'source_header': header,
                })
        result['dimensions'] = dimensions
        if format_name != 'combination':
            return
        if price_header not in headers:
            result['needs_mapping'] = True
            result['errors'].append('组合价格表需要指定价格列')
            return
        code_header = canonical.get('code')
        label_header = canonical.get('label')
        default_header = canonical.get('is_default')
        seen_signatures = set()
        for row_no, row in enumerate(rows[1:], start=2):
            data = dict(zip(headers, row))
            if all(cls._text(data.get(header)) == '' for header in dimension_headers):
                continue
            try:
                price = Decimal(str(data.get(price_header)))
            except Exception:
                result['errors'].append(f'{ws.title} 第{row_no}行: 价格无效')
                continue
            selections = {}
            for dimension in dimensions:
                value = cls._text(data.get(dimension['source_header']))
                if value:
                    selections[dimension['dimension_key']] = value
            if len(selections) != len(dimensions):
                result['errors'].append(f'{ws.title} 第{row_no}行: 配置组合不完整')
                continue
            signature = ProductPriceMatrix.build_signature(selections)
            if signature in seen_signatures:
                result['errors'].append(f'{ws.title} 第{row_no}行: 配置组合重复')
                continue
            seen_signatures.add(signature)
            result['price_entries'].append({'config': selections, 'price': price})
            code = cls._text(data.get(code_header)) if code_header else f'COMBO-{row_no}'
            label = cls._text(data.get(label_header)) if label_header else code
            result['presets'].append({
                'code': code or f'COMBO-{row_no}', 'label': label or code,
                'selections': selections,
                'is_default': cls._truthy(data.get(default_header)) if default_header else False,
            })

    @classmethod
    def _parse_price_sheet(cls, ws, result, dimensions):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        headers = [cls._text(value) for value in rows[0]]
        canonical = {cls._canonical_header(header): header for header in headers if cls._canonical_header(header)}
        if result['pricing_mode'] == 'MATRIX':
            price_header = canonical.get('price')
            if not price_header:
                result['errors'].append(f'{ws.title}: 缺少 price/最终价格列')
                return
            unknown = [header for header in headers if header and header != price_header and header not in dimensions]
            if unknown:
                result['errors'].append(f'{ws.title}: 未知维度列：{"、".join(unknown)}')
                return
            for row_no, row in enumerate(rows[1:], start=2):
                data = dict(zip(headers, row))
                if not any(cls._text(value) for value in row):
                    continue
                try:
                    price = Decimal(str(data.get(price_header)))
                except Exception:
                    result['errors'].append(f'{ws.title} 第{row_no}行: 价格无效')
                    continue
                config = {key: cls._text(data.get(key)) for key in dimensions if cls._text(data.get(key))}
                if not config:
                    result['errors'].append(f'{ws.title} 第{row_no}行: 未填写配置选项')
                    continue
                result['price_entries'].append({'config': config, 'price': price})
        else:
            required = ('dimension_key', 'option_key', 'price_delta')
            if any(name not in canonical for name in required):
                result['errors'].append(f'{ws.title}: 缺少 dimension_key、option_key 或 price_delta 列')
                return
            for row_no, row in enumerate(rows[1:], start=2):
                data = dict(zip(headers, row))
                try:
                    result['price_entries'].append({
                        'dimension_key': cls._text(data.get(canonical['dimension_key'])),
                        'option_key': cls._text(data.get(canonical['option_key'])),
                        'price_delta': Decimal(str(data.get(canonical['price_delta']))),
                    })
                except Exception:
                    result['errors'].append(f'{ws.title} 第{row_no}行: 加价格式错误')

    @classmethod
    def _parse_preset_sheet(cls, ws, result, dimensions):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        headers = [cls._text(value) for value in rows[0]]
        canonical = {cls._canonical_header(header): header for header in headers if cls._canonical_header(header)}
        for row_no, row in enumerate(rows[1:], start=2):
            data = dict(zip(headers, row))
            selections = {key: cls._text(data.get(key)) for key in dimensions if cls._text(data.get(key))}
            if not selections:
                continue
            result['presets'].append({
                'code': cls._text(data.get(canonical.get('code'))) or f'PRESET-{row_no}',
                'label': cls._text(data.get(canonical.get('label'))),
                'selections': selections,
                'is_default': cls._truthy(data.get(canonical.get('is_default'))),
            })

    @classmethod
    def _finish(cls, product, result):
        result['success_count'] = len(result['price_entries'])
        if not result['needs_mapping'] and not result['dimensions']:
            result['errors'].append('未解析到任何配置维度，禁止导入')
        defaults = [item for item in result['presets'] if item.get('is_default')]
        if len(defaults) > 1:
            result['errors'].append('默认配置不能超过一条')
        result['failed_count'] = len(result['errors'])
        if not result['price_entries']:
            result['warnings'].append('文件未包含完整组合价格；本次只合并维度和选项，现有定价与默认配置将保留')
        result['impact'] = {
            'existing_dimensions': product.config_dimensions.count(),
            'existing_matrix_rows': product.price_matrix.count(),
            'existing_rules': product.price_rules.count(),
            'existing_presets': product.config_presets.count(),
            'incoming_dimensions': len(result['dimensions']),
            'incoming_prices': len(result['price_entries']),
            'incoming_presets': len(result['presets']),
        }
        return result

    @classmethod
    @transaction.atomic
    def execute_import(cls, product, parsed, replace_dimensions=False, replace_prices=True):
        if parsed.get('needs_mapping') or parsed.get('errors') or not parsed.get('dimensions'):
            raise ValueError('解析结果不完整，不能执行导入')
        has_prices = bool(parsed.get('price_entries'))
        if replace_dimensions and not has_prices and (
                product.price_matrix.exists() or product.price_rules.exists() or product.config_presets.exists()):
            raise ValueError('完全替换维度必须同时提供完整价格数据；请改用默认的合并模式')
        if replace_dimensions:
            product.config_dimensions.all().delete()
        for item in parsed['dimensions']:
            defaults = {
                'dimension_label': item['dimension_label'], 'options': item['options'],
                'parent_dimension': item.get('parent_dimension', ''),
                'is_required': item.get('is_required', True),
                'sort_order': item.get('sort_order', 0),
            }
            ProductConfigDimension.objects.update_or_create(
                product=product, dimension_key=item['dimension_key'], defaults=defaults)
        if has_prices and replace_prices:
            product.price_matrix.all().delete()
            product.price_rules.all().delete()
            product.config_presets.all().delete()
            if parsed['pricing_mode'] == 'MATRIX':
                for entry in parsed['price_entries']:
                    config = ProductPriceMatrix.normalize_selections(entry['config'])
                    ProductPriceMatrix.objects.create(
                        product=product,
                        config_signature=ProductPriceMatrix.build_signature(config),
                        config_attributes=config, price=entry['price'])
                product.min_price = min(entry['price'] for entry in parsed['price_entries'])
            else:
                for order, entry in enumerate(parsed['price_entries']):
                    ProductPriceRule.objects.create(product=product, sort_order=order, **entry)
            for order, preset in enumerate(parsed.get('presets', [])):
                ProductConfigPreset.objects.create(
                    product=product, code=preset['code'], label=preset.get('label', ''),
                    selections=ProductPriceMatrix.normalize_selections(preset['selections']),
                    is_default=preset.get('is_default', False), sort_order=order)
            product.pricing_mode = parsed['pricing_mode']
            if parsed.get('base_price') is not None:
                product.base_price = parsed['base_price']
            product.save(update_fields=['pricing_mode', 'base_price', 'min_price'])
        return {
            'dimensions': len(parsed['dimensions']), 'prices': len(parsed.get('price_entries', [])),
            'presets': len(parsed.get('presets', [])), 'mode': 'replace' if replace_dimensions else 'merge',
        }


class SafeConfigExportService:
    """导出可直接回导的完整标准配置工作簿。"""

    @staticmethod
    def export(product):
        workbook = openpyxl.Workbook()
        dimensions = workbook.active
        dimensions.title = 'dimensions'
        dimensions.append(['dimension_key', 'dimension_label', 'options', 'parent_dimension', 'is_required', 'sort_order'])
        dimension_keys = []
        for item in product.config_dimensions.all():
            dimension_keys.append(item.dimension_key)
            options = ','.join(
                f'{option.get("key")}|{option.get("label")}'
                if option.get('label') and option.get('label') != option.get('key') else str(option.get('key', ''))
                for option in item.options or [] if isinstance(option, dict))
            dimensions.append([item.dimension_key, item.dimension_label, options, item.parent_dimension,
                               'TRUE' if item.is_required else 'FALSE', item.sort_order])
        mode = workbook.create_sheet('pricing_mode')
        mode.append(['mode', 'base_price'])
        mode.append([product.pricing_mode, product.base_price])
        if product.pricing_mode == 'MATRIX':
            prices = workbook.create_sheet('matrix')
            prices.append([*dimension_keys, 'price'])
            for item in product.price_matrix.all():
                prices.append([*(item.config_attributes.get(key, '') for key in dimension_keys), item.price])
        else:
            prices = workbook.create_sheet('rules')
            prices.append(['dimension_key', 'option_key', 'price_delta'])
            for item in product.price_rules.all():
                prices.append([item.dimension_key, item.option_key, item.price_delta])
        presets = workbook.create_sheet('presets')
        presets.append(['code', 'label', 'is_default', *dimension_keys])
        for item in product.config_presets.all():
            presets.append([item.code, item.label, 'TRUE' if item.is_default else 'FALSE',
                            *(item.selections.get(key, '') for key in dimension_keys)])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
