"""Quote services."""
import os
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from django.conf import settings
from django.db import transaction
from django.template.loader import render_to_string
from rest_framework.exceptions import ValidationError

try:
    from weasyprint import HTML
    HAS_WEASYPRINT = True
except (ImportError, OSError):
    HAS_WEASYPRINT = False

from .models import Quote, QuoteItem

VALID_TRANSITIONS = {
    'DRAFT': ['SENT', 'CANCELLED'],
    'SENT': ['CONFIRMED', 'CANCELLED'],
    'CONFIRMED': ['CANCELLED'],
    'CANCELLED': [],
}


class QuoteService:
    @staticmethod
    def validate_status_change(current: str, new_status: str) -> bool:
        return new_status in VALID_TRANSITIONS.get(current, [])

    @staticmethod
    @transaction.atomic
    def duplicate(quote_id: int, user) -> Quote:
        original = Quote.objects.prefetch_related('items').get(pk=quote_id)
        new_quote = Quote.objects.create(
            title=f"{original.title}(副本)",
            customer_name=original.customer_name,
            status='DRAFT',
            notes=original.notes,
            terms=original.terms,
            discount=original.discount,
            created_by=user,
        )
        for item in original.items.all():
            QuoteItem.objects.create(
                quote=new_quote,
                product=item.product,
                product_name=item.product_name,
                config_name=item.config_name,
                config_attributes=item.config_attributes,
                image=item.image,
                image_url=item.image_url,
                unit_price=item.unit_price,
                quantity=item.quantity,
                discount=Decimal('0'),
                sort_order=item.sort_order,
            )
        new_quote.recalculate_total()
        return new_quote

    @staticmethod
    def export_pdf(quote_id: int) -> bytes:
        if not HAS_WEASYPRINT:
            raise ImportError('WeasyPrint is not installed. Install it with: pip install WeasyPrint')
        quote = Quote.objects.prefetch_related('items', 'items__image').get(pk=quote_id)
        media_root = str(settings.MEDIA_ROOT)
        # 把 image_url 解析成模板可直接 src 引用的绝对路径
        items_with_abs_image = []
        for item in quote.items.all():
            abs_path = ''
            if item.image_url:
                candidate = os.path.join(media_root, item.image_url)
                if os.path.exists(candidate):
                    # 用 file:// 让 WeasyPrint 把它当本地资源加载
                    abs_path = 'file:///' + candidate.replace('\\', '/').lstrip('/')
            items_with_abs_image.append({'item': item, 'abs_image': abs_path})
        html_string = render_to_string('quotes/pdf_template.html', {
            'quote': quote,
            'items_with_abs_image': items_with_abs_image,
            'MEDIA_ROOT': media_root,
        })
        # base_url 让相对路径能解析到 MEDIA_ROOT
        pdf = HTML(string=html_string, base_url=media_root).write_pdf()
        return pdf

    @staticmethod
    def _resolve_image(product, image_id=None):
        from products.models import ProductImage
        if image_id:
            image = ProductImage.objects.filter(pk=image_id, product=product).first()
            if image:
                return image
        return (ProductImage.objects.filter(product=product, is_cover=True).first()
                or ProductImage.objects.filter(product=product).first())

    @staticmethod
    @transaction.atomic
    def add_item_from_product(quote: Quote, product, selections: dict,
                              image_id=None, quantity=1):
        """一键加入报价单：算价、默认封面、中文配置摘要和总价重算。"""
        from products.services import PriceCalculationService
        from products.models import ProductPriceMatrix

        selections = ProductPriceMatrix.normalize_selections(selections)
        result = PriceCalculationService.calculate(product, selections)
        if not result['valid']:
            raise ValidationError({
                'detail': result.get('reason', '配置无效'),
                'missing_dimensions': result.get('missing_dimensions', []),
                'invalid_selections': result.get('invalid_selections', []),
            })
        image = QuoteService._resolve_image(product, image_id)
        item = QuoteItem.objects.create(
            quote=quote, product=product, product_name=product.name,
            config_name=QuoteService._summarize_selections(product, selections),
            config_attributes=selections,
            unit_price=Decimal(str(result['price'])), quantity=quantity,
            discount=Decimal('0'), image=image,
            image_url=image.image_path if image else '', sort_order=quote.items.count(),
        )
        quote.recalculate_total()
        return item

    @staticmethod
    @transaction.atomic
    def update_item_from_product(item: QuoteItem, product, selections: dict,
                                 image_id=None, quantity=None):
        """改配置时原位更新 QuoteItem，避免产生重复明细。"""
        from products.services import PriceCalculationService
        from products.models import ProductPriceMatrix

        selections = ProductPriceMatrix.normalize_selections(selections)
        result = PriceCalculationService.calculate(product, selections)
        if not result['valid']:
            raise ValidationError({
                'detail': result.get('reason', '配置无效'),
                'missing_dimensions': result.get('missing_dimensions', []),
                'invalid_selections': result.get('invalid_selections', []),
            })
        image = QuoteService._resolve_image(product, image_id) if image_id is not None else item.image
        if image is None:
            image = QuoteService._resolve_image(product)
        item.product = product
        item.product_name = product.name
        item.config_name = QuoteService._summarize_selections(product, selections)
        item.config_attributes = selections
        item.unit_price = Decimal(str(result['price']))
        if quantity is not None:
            item.quantity = quantity
        item.image = image
        item.image_url = image.image_path if image else ''
        item.save()
        item.quote.recalculate_total()
        return item

    @staticmethod
    def _summarize_selections(product, selections: dict) -> str:
        """使用维度和选项的展示名生成可读配置摘要。"""
        from products.models import ProductConfigDimension
        dims = ProductConfigDimension.objects.filter(product=product)
        dim_map = {d.dimension_key: d for d in dims}
        parts = []
        for key, value in selections.items():
            dim = dim_map.get(key)
            dim_label = dim.dimension_label if dim else key
            option_label = value
            if dim:
                for option in dim.options or []:
                    if isinstance(option, dict) and str(option.get('key')) == str(value):
                        option_label = option.get('label') or value
                        break
            parts.append(f'{dim_label}:{option_label}')
        return ' / '.join(parts)



# ─── QT-3 Excel 导出（两种格式） ─────────────────────────────────────────────

class QuoteExcelService:
    """导出报价单 Excel：两种格式的表头不同，明细列一致。

    列（明细统一）：序号 | 产品名称 | 产品描述 | 配置(默认/自定义) | 颜色 | 单价 | 数量 | 小计 | 产地 | 货期 | 品牌 | 图片
    """
    ORIGIN_MAP = {'IMPORT': '进口', 'DOMESTIC': '国产'}
    LEAD_TIME_MAP = {
        'WITHIN_45D': '45天内', '2_4M_VIETNAM': '2-4月【越南】',
        '2_4M_MALAYSIA': '2-4月【马来西亚】', '4_6M_EU': '4-6月【荷兰/意大利/德国】',
    }

    ITEM_HEADERS = ['序号', '产品名称', '产品描述', '配置', '颜色',
                    '单价', '数量', '小计', '产地', '货期', '品牌', '图片']

    @classmethod
    def _extract_color(cls, item) -> str:
        attrs = item.config_attributes or {}
        for k in ('color', '颜色', 'frame_color_back', 'frame_color', '框架颜色', '框架颜色（背框）'):
            if k in attrs:
                return str(attrs[k])
        return ''

    @classmethod
    def _config_summary(cls, item) -> str:
        return item.config_name or ''

    @classmethod
    def _fill_items(cls, ws, quote, start_row: int):
        for header_idx, h in enumerate(cls.ITEM_HEADERS, start=1):
            ws.cell(row=start_row, column=header_idx, value=h)
        ws.column_dimensions['L'].width = 14
        r = start_row + 1
        for i, item in enumerate(quote.items.all(), start=1):
            p = item.product
            row = [
                i, item.product_name, (p.description if p else ''),
                cls._config_summary(item), cls._extract_color(item),
                float(item.unit_price), item.quantity, float(item.subtotal),
                cls.ORIGIN_MAP.get(p.origin, '') if p else '',
                cls.LEAD_TIME_MAP.get(p.lead_time, '') if p else '',
                (p.brand.name if p and p.brand else ''), '',
            ]
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=r, column=col_idx, value=value)
            if item.image_url:
                image_path = os.path.join(str(settings.MEDIA_ROOT), item.image_url)
                if os.path.exists(image_path):
                    try:
                        image = OpenpyxlImage(image_path)
                        image.width, image.height = 72, 54
                        ws.add_image(image, f'L{r}')
                        ws.row_dimensions[r].height = 44
                    except (OSError, ValueError):
                        ws.cell(row=r, column=12, value='图片不可用')
            r += 1
        subtotal = sum((item.subtotal for item in quote.items.all()), Decimal('0'))
        discount_amount = subtotal - quote.total_amount
        summary_rows = [
            ('折前合计', subtotal),
            ('整单折扣', f'{quote.discount}%'),
            ('折扣金额', discount_amount),
            ('折后总额', quote.total_amount),
        ]
        for label, value in summary_rows:
            ws.cell(row=r, column=7, value=label)
            ws.cell(row=r, column=8, value=float(value) if isinstance(value, Decimal) else value)
            r += 1
        return r

    @classmethod
    def export_sales_order(cls, quote) -> bytes:
        """格式一：销售订单（image5.png 风格）"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '销售订单'
        ws.append(['销售订单'])
        ws.append([f'合同编号：ZKZY{quote.id:08d}'])
        ws.append(['买方：', quote.customer_name, '', '', '卖方：', '杭州智楷家具有限公司'])
        ws.append(['地址：', '', '', '', '地址：', '杭州市拱墅区'])
        ws.append(['联系人：', '', '', '', '联系人：', quote.created_by.display_name if quote.created_by else ''])
        ws.append(['电话：', '', '', '', '电话：', ''])
        ws.append(['交货日期：', '详见清单备注', '', '', '付款方式：', '汇款'])
        ws.append([''])
        cls._fill_items(ws, quote, start_row=9)
        if quote.notes:
            ws.append(['备注：', quote.notes])
        if quote.terms:
            ws.append(['条款：', quote.terms])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    @classmethod
    def export_quotation(cls, quote) -> bytes:
        """格式二：报价单（image6.png 风格，中英双语抬头）"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '报价单'
        ws.append(['报价单 / Quotation'])
        ws.append(['ZhiKai Furniture, Hangzhou'])
        ws.append(['杭州市拱墅区'])
        ws.append([''])
        ws.append(['Project / 项目名称：', quote.title,
                   '', '', 'Quote date / 报价日期：', quote.created_at.strftime('%Y-%m-%d')])
        ws.append(['Pricing Currency / 计价币种：', 'RMB (yuan) / 人民币（元）',
                   '', '', 'Account Manager / 销售：', quote.created_by.display_name if quote.created_by else ''])
        ws.append(['Company / 公司：', '杭州智楷家具有限公司',
                   '', '', 'Telephone / 电话：', ''])
        ws.append([''])
        cls._fill_items(ws, quote, start_row=9)
        if quote.notes:
            ws.append(['Notes / 备注：', quote.notes])
        if quote.terms:
            ws.append(['Terms / 条款：', quote.terms])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
