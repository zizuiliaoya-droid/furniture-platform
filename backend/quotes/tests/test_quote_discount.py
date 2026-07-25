"""QT-6：整单折扣改造回归测试。"""
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from rest_framework.test import APIClient

from quotes.models import Quote, QuoteItem
from quotes.services import QuoteExcelService


@pytest.fixture
def api(admin_user):
    client = APIClient()
    client.force_authenticate(user=admin_user)
    return client


@pytest.mark.django_db
class TestQuoteDiscount:
    def test_whole_quote_discount_applies(self, api, quote, product_matrix):
        # 加两条明细：2580 * 2 = 5160
        for _ in range(2):
            QuoteItem.objects.create(
                quote=quote, product=product_matrix, product_name=product_matrix.name,
                unit_price=Decimal('2580'), quantity=1, discount=Decimal('0'),
            )
        quote.recalculate_total()
        assert quote.total_amount == Decimal('5160.00')

        # 设置整单 10% 折扣 → 5160 * 0.9 = 4644
        resp = api.patch(f'/api/quotes/{quote.id}/', {'discount': '10'}, format='json')
        assert resp.status_code == 200, resp.content
        quote.refresh_from_db()
        assert quote.total_amount == Decimal('4644.00')

    def test_zero_discount_equals_subtotal(self, api, quote, product_matrix):
        QuoteItem.objects.create(
            quote=quote, product=product_matrix, product_name=product_matrix.name,
            unit_price=Decimal('1000'), quantity=3, discount=Decimal('0'),
        )
        quote.recalculate_total()
        assert quote.total_amount == Decimal('3000.00')

    def test_duplicate_preserves_whole_quote_discount_and_clears_item_discount(
            self, api, quote, product_matrix):
        quote.discount = Decimal('10')
        quote.save(update_fields=['discount'])
        QuoteItem.objects.create(
            quote=quote, product=product_matrix, product_name=product_matrix.name,
            unit_price=Decimal('1000'), quantity=2, discount=Decimal('5'),
        )
        quote.recalculate_total()

        response = api.post(f'/api/quotes/{quote.id}/duplicate/')
        assert response.status_code == 201, response.content
        duplicate = Quote.objects.get(pk=response.data['id'])
        assert duplicate.discount == Decimal('10')
        assert duplicate.items.get().discount == Decimal('0')
        assert duplicate.total_amount == Decimal('1800.00')

    def test_excel_shows_before_discount_discount_amount_and_final_total(
            self, quote, product_matrix):
        quote.discount = Decimal('10')
        quote.save(update_fields=['discount'])
        QuoteItem.objects.create(
            quote=quote, product=product_matrix, product_name=product_matrix.name,
            unit_price=Decimal('1000'), quantity=2, discount=Decimal('0'),
        )
        quote.recalculate_total()

        workbook = openpyxl.load_workbook(BytesIO(QuoteExcelService.export_quotation(quote)))
        sheet = workbook.active
        summary = {
            sheet.cell(row=row, column=7).value: sheet.cell(row=row, column=8).value
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=7).value
        }
        assert summary['折前合计'] == 2000
        assert summary['整单折扣'] == '10%'
        assert summary['折扣金额'] == 200
        assert summary['折后总额'] == 1800