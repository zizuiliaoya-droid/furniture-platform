"""Agent Excel 导入必须预览、确认且防重放。"""
from io import BytesIO

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.authtoken.models import Token
from rest_framework.test import APIClient


def agent_client(user):
    token = Token.objects.create(user=user)
    client = APIClient()
    client.credentials(
        HTTP_AUTHORIZATION=f'Token {token.key}',
        HTTP_X_AGENT_SKILL='furniture-import',
    )
    return client


def config_workbook(price=2800):
    workbook = openpyxl.Workbook()
    dimensions = workbook.active
    dimensions.title = 'dimensions'
    dimensions.append([
        'dimension_key', 'dimension_label', 'options',
        'parent_dimension', 'is_required', 'sort_order',
    ])
    dimensions.append(['color', '颜色', 'red|红,blue|蓝', '', 'TRUE', 1])
    dimensions.append(['size', '尺寸', 'S,L', '', 'TRUE', 2])
    mode = workbook.create_sheet('pricing_mode')
    mode.append(['mode', 'base_price'])
    mode.append(['MATRIX', ''])
    matrix = workbook.create_sheet('matrix')
    matrix.append(['color', 'size', 'price'])
    matrix.append(['red', 'L', price])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def upload(content, name='config.xlsx'):
    return SimpleUploadedFile(
        name,
        content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@pytest.mark.django_db
class TestAgentImportWorkflow:
    def test_preview_does_not_mutate_and_confirm_is_one_time(
        self,
        admin_user,
        product_matrix,
    ):
        content = config_workbook()
        client = agent_client(admin_user)
        old_price = product_matrix.price_matrix.get().price

        preview = client.post(
            f'/api/agent/products/{product_matrix.id}/config-import/preview/',
            {'file': upload(content)},
            format='multipart',
        )

        assert preview.status_code == 200
        assert preview.data['confirmation_token']
        assert preview.data['preview']['errors'] == []
        assert product_matrix.price_matrix.get().price == old_price

        confirm_data = {
            'file': upload(content),
            'confirmation_token': preview.data['confirmation_token'],
            'replace_prices': 'true',
            'replace_dimensions': 'false',
        }
        confirmed = client.post(
            f'/api/agent/products/{product_matrix.id}/config-import/confirm/',
            confirm_data,
            format='multipart',
        )
        replay = client.post(
            f'/api/agent/products/{product_matrix.id}/config-import/confirm/',
            {
                'file': upload(content),
                'confirmation_token': preview.data['confirmation_token'],
                'replace_prices': 'true',
                'replace_dimensions': 'false',
            },
            format='multipart',
        )

        assert confirmed.status_code == 200
        assert replay.status_code == 409
        product_matrix.refresh_from_db()
        assert str(product_matrix.price_matrix.get().price) == '2800.00'

    def test_confirmation_is_bound_to_exact_file(self, admin_user, product_matrix):
        original = config_workbook(2800)
        client = agent_client(admin_user)
        preview = client.post(
            f'/api/agent/products/{product_matrix.id}/config-import/preview/',
            {'file': upload(original)},
            format='multipart',
        )

        response = client.post(
            f'/api/agent/products/{product_matrix.id}/config-import/confirm/',
            {
                'file': upload(config_workbook(9999)),
                'confirmation_token': preview.data['confirmation_token'],
                'replace_prices': 'true',
                'replace_dimensions': 'false',
            },
            format='multipart',
        )

        assert response.status_code == 400
        assert str(product_matrix.price_matrix.get().price) == '2580.00'

    def test_confirm_without_preview_token_is_rejected(self, admin_user, product_matrix):
        response = agent_client(admin_user).post(
            f'/api/agent/products/{product_matrix.id}/config-import/confirm/',
            {'file': upload(config_workbook())},
            format='multipart',
        )

        assert response.status_code == 400
        assert str(product_matrix.price_matrix.get().price) == '2580.00'
